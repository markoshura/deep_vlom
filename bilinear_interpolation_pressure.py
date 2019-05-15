
from openpyxl import load_workbook
from math import log10

# Load in the workbook
wb = load_workbook('C:/Users/xaha9/Desktop/ШР/deep_lom/aluminium.xlsx')


sheet = wb.get_sheet_by_name('pressure_ferrum_исходник')



def bilinear_interpolation_pressure(T, rho):
    x1 = 0
    x2 = 0
    y1 = 0
    y2 = 0

    for cellObj in sheet['B1':'OL1']:
        for cell in cellObj:
            if cell.value <= log10(rho):
                x1 = cell
            else:

                break


    x2 = sheet.cell(row=x1.row, column=x1.column + 1)

    #print("x1 = ", x1, "X2 = ", x2)
    for cellObj in sheet['A2':'A502']:
        for cell in cellObj:
            if cell.value >= log10(T):
                y1 = cell
            else:
                break

    y2 = sheet.cell(row=y1.row + 1, column=y1.column)
    #print("y1 = ", y1, "y2 = ", y2)


    def interpolate(x1, x2, y1, y2, T, rho):

        f1 = (10 **(x2.value) - rho) / (10 **(x2.value) - 10 **(x1.value)) * sheet.cell(row=y1.row, column=x1.column).value + (rho - 10 **(x1.value)) / (10 **(x2.value) - 10 **(x1.value)) * sheet.cell(row=y1.row, column=x2.column).value
        f2 = (10 **(x2.value) - rho) / (10 **(x2.value) - 10 **(x1.value)) * sheet.cell(row=y2.row, column=x1.column).value + (rho - 10 **(x1.value)) / (10 **(x2.value) - 10 **(x1.value)) * sheet.cell(row=y2.row, column=x2.column).value

        return (10 **(y2.value) - T) / (10 **(y2.value) - 10 **(y1.value)) * f1 + (T - 10 **(y1.value)) / (10 **(y2.value) - 10 **(y1.value)) * f2

    return interpolate(x1, x2, y1, y2, T, rho)

print(bilinear_interpolation_pressure(100, 0.01))




